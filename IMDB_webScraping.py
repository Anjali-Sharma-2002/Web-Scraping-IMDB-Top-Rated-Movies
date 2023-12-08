import requests
from bs4 import BeautifulSoup
import openpyxl

# create the excel file 

excel = openpyxl.Workbook()
print(excel.sheetnames)
sheet = excel.active
sheet.title = "Top Rated movies"
print(excel.sheetnames)

sheet.append(['Movie Rank', 'Movie Name', 'Release Year', 'IMDB Rating', 'Votes Count'])


try:
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
    }
    source = requests.get("https://www.imdb.com/chart/top/", headers=headers)
    source.raise_for_status()   # to get error if url not right

    soup = BeautifulSoup(source.text, 'html.parser')
    movies = soup.find('ul', class_ = "ipc-metadata-list ipc-metadata-list--dividers-between sc-9d2f6de0-0 iMNUXk compact-list-view ipc-metadata-list--base")
    movies = movies.find_all('li', class_ = "ipc-metadata-list-summary-item sc-59b6048d-0 cuaJSp cli-parent")


    for movie in movies:
        rank = movie.find('a', class_ = "ipc-title-link-wrapper").text.split('.')[0]
        name = movie.find('a', class_ = "ipc-title-link-wrapper").text.split('.')[1]
        year = movie.find('span', class_ = "sc-479faa3c-8 bNrEFi cli-title-metadata-item").text
        rating = movie.find('span', class_ = "ipc-rating-star ipc-rating-star--base ipc-rating-star--imdb ratingGroup--imdb-rating").text.split('(')[0]
        votes = movie.find('span', class_ = "ipc-rating-star ipc-rating-star--base ipc-rating-star--imdb ratingGroup--imdb-rating").text.split('(')[1].split(')')[0]
        print(rank, name, year, rating,votes)
        sheet.append([rank, name, year, rating,votes])

except Exception as e:
    print(e)



excel.save('IMDB_rating.xlsx')
